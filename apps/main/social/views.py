from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.http import JsonResponse, HttpResponseForbidden
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.db.models import Q, Count
from django.core.paginator import Paginator
from django.utils.translation import gettext as _
from django.contrib.auth import get_user_model
from django.views.decorators.http import require_POST, require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.utils import timezone
from datetime import datetime, timedelta
import re

from .models import Post, Comment, Like, Follow, UserProfile, Share, Hashtag, PostHashtag, CommentLike, Report, ModerationAction, ContentFilter, ModerationLog
from .forms import PostForm, CommentForm, UserProfileForm, SearchForm, ShareForm, ReactionForm, HashtagForm, ReportForm, SearchReportForm, BulkModerationForm, ModerationActionForm, ContentFilterForm

User = get_user_model()


@login_required
def feed(request):
    """Feed principal da rede social"""
    # Buscar posts de usuários que o usuário segue + posts públicos + posts próprios
    following_users = request.user.following.values_list('following_id', flat=True)
    
    # Filtrar posts baseado em permissões de moderação
    if request.user.is_superuser or request.user.is_staff or request.user.has_perm('social.can_moderate_content'):
        # Moderadores veem todos os posts, incluindo os ocultos
        posts = Post.objects.filter(
            Q(author__in=following_users) | Q(is_public=True) | Q(author=request.user)
        ).select_related('author').prefetch_related(
            'likes', 'comments', 'hashtags', 'reports'
        ).order_by('-created_at')
    else:
        # Usuários regulares não veem posts ocultos ou deletados
        hidden_posts = ModerationAction.objects.filter(
            action_type__in=['hide_content', 'delete_content'],
            is_active=True,
            target_post__isnull=False
        ).values_list('target_post_id', flat=True)
        
        posts = Post.objects.filter(
            Q(author__in=following_users) | Q(is_public=True) | Q(author=request.user)
        ).exclude(
            id__in=hidden_posts
        ).select_related('author').prefetch_related(
            'likes', 'comments', 'hashtags', 'reports'
        ).order_by('-created_at')
    
    # Anotar posts com informação se o usuário atual deu like e status de moderação
    for post in posts:
        post.is_liked_by_current_user = post.is_liked_by(request.user)
        
        # Obter a reação atual do usuário
        try:
            user_like = post.likes.get(user=request.user)
            post.current_user_reaction = user_like.reaction_type
        except Like.DoesNotExist:
            post.current_user_reaction = None
            
        # Verificar se o post foi denunciado
        post.is_flagged = post.reports.filter(status='pending').exists()
        # Verificar se o post está oculto por moderação
        post.is_hidden = ModerationAction.objects.filter(
            target_post=post,
            action_type__in=['hide_content', 'delete_content'],
            is_active=True
        ).exists()
    
    # Paginação
    paginator = Paginator(posts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Buscar perfil do usuário
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    # Atualizar estatísticas do perfil (garantir que estejam atualizadas)
    profile.update_statistics()
    
    # Estatísticas do usuário
    user_stats = {
        'posts_count': profile.total_posts,
        'followers_count': request.user.followers.count(),
        'following_count': request.user.following.count(),
        'total_likes_received': profile.total_likes_received,
        'total_comments_received': profile.total_comments_received,
    }
    
    # Hashtags populares
    popular_hashtags = Hashtag.objects.filter(posts_count__gt=0).order_by('-posts_count')[:10]
    
    # Usuários sugeridos (não seguidos pelo usuário atual)
    following_users = request.user.following.values_list('following_id', flat=True)
    suggested_users = User.objects.exclude(
        id__in=list(following_users) + [request.user.id]
    ).annotate(
        posts_count=Count('social_posts')
    ).filter(posts_count__gt=0).order_by('-posts_count')[:10]
    
    # Estatísticas da rede
    from datetime import date
    today = date.today()
    network_stats = {
        'total_users': User.objects.count(),
        'total_posts': Post.objects.count(),
        'posts_today': Post.objects.filter(created_at__date=today).count(),
    }
    
    # Verificar permissões de moderação
    can_moderate = request.user.is_superuser or request.user.is_staff or request.user.has_perm('social.can_moderate_content')
    
    # Estatísticas de moderação (apenas para moderadores)
    moderation_stats = {}
    if can_moderate:
        moderation_stats = {
            'pending_reports': Report.objects.filter(status='pending').count(),
            'total_reports': Report.objects.count(),
            'reports_today': Report.objects.filter(created_at__date=today).count(),
            'active_filters': ContentFilter.objects.filter(is_active=True).count(),
        }
    
    # Formulário para criar novo post
    if request.method == 'POST':
        form = PostForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                post = form.save(commit=False)
                post.author = request.user
                
                # Processar hashtags
                hashtags = form.cleaned_data.get('hashtags', [])
                
                post.save()
                
                # Adicionar hashtags ao post
                for hashtag_name in hashtags:
                    hashtag, created = Hashtag.objects.get_or_create(name=hashtag_name)
                    PostHashtag.objects.create(post=post, hashtag=hashtag)
                    hashtag.update_posts_count()
                
                messages.success(request, _('Post criado com sucesso!'))
                return redirect('social:feed')
            except Exception as e:
                messages.error(request, _('Erro ao criar post. Tente novamente.'))
                # Log do erro para debug
                print(f"Erro ao criar post: {e}")
        else:
            # Exibir erros de validação
            for field, errors in form.errors.items():
                for error in errors:
                    messages.error(request, f"Erro no campo {field}: {error}")
    else:
        form = PostForm()
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'profile': profile,
        'user_stats': user_stats,
        'popular_hashtags': popular_hashtags,
        'suggested_users': suggested_users,
        'network_stats': network_stats,
        'moderation_stats': moderation_stats,
        'can_moderate': can_moderate,
        'segment': 'feed',
        'parent': 'social',
    }
    return render(request, 'social/feed.html', context)


@login_required
def my_posts(request):
    """Posts do usuário logado"""
    posts = Post.objects.filter(author=request.user).select_related('author').prefetch_related(
        'likes', 'comments', 'hashtags', 'reports'
    ).order_by('-is_pinned', '-created_at')
    
    # Verificar se o usuário pode moderar
    can_moderate = request.user.is_superuser or request.user.is_staff or request.user.has_perm('social.can_moderate_content')
    
    # Anotar posts com informação se o usuário atual deu like e status de moderação
    for post in posts:
        post.is_liked_by_current_user = post.is_liked_by(request.user)
        # Verificar reação atual do usuário
        user_like = post.likes.filter(user=request.user).first()
        post.current_user_reaction = user_like.reaction_type if user_like else None
        
        # Status de moderação (apenas para moderadores)
        if can_moderate:
            post.is_flagged = post.reports.filter(status='pending').exists()
            post.is_hidden = ModerationAction.objects.filter(
                action_type='hide_content',
                target_post=post,
                is_active=True
            ).exists()
        else:
            post.is_flagged = False
            post.is_hidden = False
    
    paginator = Paginator(posts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Timestamp para cache busting de avatares
    import time
    timestamp = int(time.time())
    
    context = {
        'page_obj': page_obj,
        'can_moderate': can_moderate,
        'timestamp': timestamp,
        'segment': 'my_posts',
        'parent': 'social',
    }
    return render(request, 'social/my_posts.html', context)


@login_required
def post_detail(request, post_id):
    """Detalhes de um post específico"""
    post = get_object_or_404(Post, id=post_id)
    
    # Verificar se o usuário pode ver o post
    if not post.is_public and post.author != request.user:
        if not Follow.objects.filter(follower=request.user, following=post.author).exists():
            messages.error(request, _('Você não tem permissão para ver este post.'))
            return redirect('social:feed')
    
    # Incrementar visualizações
    post.increment_views()
    
    # Adicionar informação se o usuário atual deu like
    post.is_liked_by_current_user = post.is_liked_by(request.user)
    
    # Formulário para comentários
    if request.method == 'POST':
        form = CommentForm(request.POST, request.FILES)
        if form.is_valid():
            comment = form.save(commit=False)
            comment.post = post
            comment.author = request.user
            comment.save()
            
            # Atualizar contador de comentários
            post.update_counts()
            
            messages.success(request, _('Comentário adicionado!'))
            return redirect('social:post_detail', post_id=post.id)
    else:
        form = CommentForm()
    
    # Buscar comentários do post
    comments = post.comments.filter(parent=None).order_by('created_at')
    
    context = {
        'post': post,
        'comments': comments,
        'form': form,
        'segment': 'post_detail',
        'parent': 'social',
    }
    return render(request, 'social/post_detail.html', context)


@login_required
@require_POST
def like_post(request, post_id):
    """Curtir/descurtir um post"""
    post = get_object_or_404(Post, id=post_id)
    
    # Verificar se o usuário pode ver o post
    if not post.is_public and post.author != request.user:
        if not Follow.objects.filter(follower=request.user, following=post.author).exists():
            return JsonResponse({'error': _('Permissão negada')}, status=403)
    
    like, created = Like.objects.get_or_create(
        post=post,
        user=request.user,
        defaults={'reaction_type': 'like'}
    )
    
    if not created:
        # Se já existe, remover a curtida
        like.delete()
        liked = False
    else:
        liked = True
    
    # Atualizar contador
    post.update_counts()
    
    return JsonResponse({
        'liked': liked,
        'likes_count': post.likes_count,
        'reaction_type': like.reaction_type if liked else None
    })


@login_required
@require_POST
def react_to_post(request, post_id):
    """Reagir a um post com diferentes emojis"""
    post = get_object_or_404(Post, id=post_id)
    form = ReactionForm(request.POST)
    
    if form.is_valid():
        reaction_type = form.cleaned_data['reaction_type']
        
        like, created = Like.objects.get_or_create(
            post=post,
            user=request.user,
            defaults={'reaction_type': reaction_type}
        )
        
        if not created:
            # Se já existe, verificar se é a mesma reação (para remover) ou trocar
            if like.reaction_type == reaction_type:
                # Mesma reação - remover
                like.delete()
                post.update_counts()
                return JsonResponse({
                    'success': True,
                    'removed': True,
                    'likes_count': post.likes_count
                })
            else:
                # Reação diferente - atualizar
                like.reaction_type = reaction_type
                like.save()
        
        # Atualizar contador
        post.update_counts()
        
        return JsonResponse({
            'success': True,
            'reaction_type': reaction_type,
            'likes_count': post.likes_count
        })
    
    return JsonResponse({'error': _('Dados inválidos')}, status=400)


@login_required
@require_POST
def like_comment(request, comment_id):
    """Curtir/descurtir um comentário"""
    comment = get_object_or_404(Comment, id=comment_id)
    
    like, created = CommentLike.objects.get_or_create(
        comment=comment,
        user=request.user
    )
    
    if not created:
        # Se já existe, remover a curtida
        like.delete()
        liked = False
    else:
        liked = True
    
    # Atualizar contador
    comment.update_likes_count()
    
    return JsonResponse({
        'liked': liked,
        'likes_count': comment.likes_count
    })


@login_required
@require_POST
def share_post(request, post_id):
    """Compartilhar um post"""
    original_post = get_object_or_404(Post, id=post_id)
    form = ShareForm(request.POST)
    
    if form.is_valid():
        # Criar novo post como compartilhamento
        post = Post.objects.create(
            author=request.user,
            content=form.cleaned_data.get('comment', ''),
            is_public=form.cleaned_data.get('is_public', True),
            link=f"/social/post/{original_post.id}/"
        )
        
        # Criar registro de compartilhamento
        Share.objects.create(
            original_post=original_post,
            user=request.user,
            comment=form.cleaned_data.get('comment', '')
        )
        
        # Atualizar contador de compartilhamentos
        original_post.shares_count += 1
        original_post.save(update_fields=['shares_count'])
        
        messages.success(request, _('Post compartilhado com sucesso!'))
        return redirect('social:feed')
    
    messages.error(request, _('Erro ao compartilhar o post.'))
    return redirect('social:post_detail', post_id=post_id)


@login_required
@require_POST
def follow_user(request, user_id):
    """Seguir/deixar de seguir um usuário"""
    try:
        # Log para debug
        print(f"Tentando seguir usuário com ID: {user_id} (tipo: {type(user_id)})")
        
        # Garantir que user_id seja um inteiro
        try:
            user_id = int(user_id)
        except (ValueError, TypeError):
            return JsonResponse({
                'error': f'ID de usuário inválido: {user_id}',
                'success': False
            }, status=400)
        
        user_to_follow = get_object_or_404(User, id=user_id)
        
        # Log adicional para debug
        print(f"Usuário encontrado: {user_to_follow.username} (ID: {user_to_follow.id})")
        print(f"Usuário atual: {request.user.username} (ID: {request.user.id})")
        
        if user_to_follow == request.user:
            return JsonResponse({'error': _('Você não pode seguir a si mesmo')}, status=400)
        
        follow, created = Follow.objects.get_or_create(
            follower=request.user,
            following=user_to_follow
        )
        
        if not created:
            # Se já existe, remover o follow
            follow.delete()
            following = False
        else:
            following = True
        
        # Calcular contadores atualizados
        followers_count = user_to_follow.followers.count()
        following_count = user_to_follow.following.count()
        
        response_data = {
            'following': following,
            'followers_count': followers_count,
            'following_count': following_count,
            'success': True
        }
        
        return JsonResponse(response_data)
        
    except Exception as e:
        return JsonResponse({
            'error': str(e),
            'success': False
        }, status=500)


@login_required
def user_profile(request, username):
    """Perfil de um usuário"""
    user = get_object_or_404(User, username=username)
    # Usar diretamente o social_profile do usuário para evitar problemas de cache
    profile = getattr(user, 'social_profile', None)
    if not profile:
        profile, created = UserProfile.objects.get_or_create(user=user)
    
    # Verificar se o usuário pode ver o perfil
    if profile.is_private and user != request.user:
        if not Follow.objects.filter(follower=request.user, following=user).exists():
            messages.error(request, _('Este perfil é privado.'))
            return redirect('social:feed')
    
    # Buscar posts do usuário
    posts = Post.objects.filter(author=user).order_by('-is_pinned', '-created_at')
    
    # Anotar posts com informação se o usuário atual deu like
    for post in posts:
        post.is_liked_by_current_user = post.is_liked_by(request.user)
    
    # Verificar se o usuário logado segue este usuário
    is_following = False
    if request.user.is_authenticated and request.user != user:
        is_following = Follow.objects.filter(
            follower=request.user,
            following=user
        ).exists()
    
    context = {
        'profile_user': user,
        'profile': profile,
        'posts': posts,
        'is_following': is_following,
        'segment': 'user_profile',
        'parent': 'social',
    }
    return render(request, 'social/user_profile.html', context)


@login_required
def edit_profile(request):
    """Editar perfil do usuário logado"""
    profile, created = UserProfile.objects.get_or_create(user=request.user)
    
    if request.method == 'POST':
        form = UserProfileForm(request.POST, request.FILES, instance=profile)
        if form.is_valid():
            form.save()
            messages.success(request, _('Perfil atualizado com sucesso!'))
            return redirect('social:user_profile', username=request.user.username)
        else:
            messages.error(request, _('Por favor, corrija os erros abaixo.'))
    else:
        form = UserProfileForm(instance=profile)

    context = {
        'form': form,
        'profile': profile,
        'segment': 'edit_profile',
        'parent': 'social',
    }
    return render(request, 'social/edit_profile.html', context)


@login_required
def search(request):
    """Buscar usuários e posts"""
    form = SearchForm(request.GET)
    results = []
    
    if form.is_valid():
        query = form.cleaned_data.get('q', '').strip()
        search_type = form.cleaned_data.get('search_type', 'all')
        date_filter = form.cleaned_data.get('date_filter', 'all')
        
        if query:
            # Aplicar filtro de data
            date_filter_obj = None
            if date_filter == 'today':
                date_filter_obj = timezone.now().date()
            elif date_filter == 'week':
                date_filter_obj = timezone.now() - timedelta(days=7)
            elif date_filter == 'month':
                date_filter_obj = timezone.now() - timedelta(days=30)
            elif date_filter == 'year':
                date_filter_obj = timezone.now() - timedelta(days=365)
            
            if search_type in ['all', 'users']:
                # Buscar usuários
                users = User.objects.filter(
                    Q(username__icontains=query) |
                    Q(first_name__icontains=query) |
                    Q(last_name__icontains=query) |
                    Q(email__icontains=query)
                ).exclude(id=request.user.id)
                
                for user in users[:10]:  # Limitar a 10 resultados
                    results.append({
                        'type': 'user',
                        'object': user,
                        'profile': getattr(user, 'social_profile', None)
                    })
            
            if search_type in ['all', 'posts']:
                # Buscar posts
                posts_query = Post.objects.filter(
                    Q(content__icontains=query) |
                    Q(author__username__icontains=query)
                ).filter(is_public=True)
                
                if date_filter_obj:
                    if isinstance(date_filter_obj, datetime):
                        posts_query = posts_query.filter(created_at__gte=date_filter_obj)
                    else:
                        posts_query = posts_query.filter(created_at__date__gte=date_filter_obj)
                
                for post in posts_query[:10]:  # Limitar a 10 resultados
                    post.is_liked_by_current_user = post.is_liked_by(request.user)
                    results.append({
                        'type': 'post',
                        'object': post
                    })
            
            if search_type in ['all', 'hashtags']:
                # Buscar hashtags
                hashtags = Hashtag.objects.filter(name__icontains=query)
                
                for hashtag in hashtags[:5]:  # Limitar a 5 resultados
                    results.append({
                        'type': 'hashtag',
                        'object': hashtag
                    })
    
    context = {
        'form': form,
        'results': results,
        'segment': 'search',
        'parent': 'social',
    }
    return render(request, 'social/search.html', context)


@login_required
def hashtag_detail(request, hashtag_name):
    """Detalhes de uma hashtag"""
    hashtag = get_object_or_404(Hashtag, name=hashtag_name.lower())
    
    # Buscar posts com esta hashtag
    posts = Post.objects.filter(
        hashtags__hashtag=hashtag,
        is_public=True
    ).order_by('-created_at')
    
    # Anotar posts com informação se o usuário atual deu like
    for post in posts:
        post.is_liked_by_current_user = post.is_liked_by(request.user)
    
    paginator = Paginator(posts, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'hashtag': hashtag,
        'page_obj': page_obj,
        'segment': 'hashtag_detail',
        'parent': 'social',
    }
    return render(request, 'social/hashtag_detail.html', context)


@login_required
def followers_list(request, username):
    """Lista de seguidores de um usuário"""
    user = get_object_or_404(User, username=username)
    followers = user.followers.all()  # Objetos Follow onde following=user
    
    paginator = Paginator(followers, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'profile_user': user,
        'page_obj': page_obj,
        'followers': page_obj,  # Adicionar para compatibilidade com template
        'segment': 'followers_list',
        'parent': 'social',
    }
    return render(request, 'social/followers_list.html', context)


@login_required
def following_list(request, username):
    """Lista de usuários que um usuário segue"""
    user = get_object_or_404(User, username=username)
    following = user.following.all()  # Objetos Follow onde follower=user
    
    paginator = Paginator(following, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'profile_user': user,
        'page_obj': page_obj,
        'following': page_obj,  # Adicionar para compatibilidade com template
        'segment': 'following_list',
        'parent': 'social',
    }
    return render(request, 'social/following_list.html', context)


@login_required
@require_POST
def pin_post(request, post_id):
    """Fixar/desfixar um post no perfil"""
    post = get_object_or_404(Post, id=post_id, author=request.user)
    
    post.is_pinned = not post.is_pinned
    post.save(update_fields=['is_pinned'])
    
    return JsonResponse({
        'pinned': post.is_pinned
    })


@login_required
@require_POST
def delete_comment(request, comment_id):
    """Deletar um comentário"""
    comment = get_object_or_404(Comment, id=comment_id)
    
    # Verificar permissão
    if comment.author != request.user and comment.post.author != request.user:
        return JsonResponse({'error': _('Permissão negada')}, status=403)
    
    post = comment.post
    comment.delete()
    
    # Atualizar contador
    post.update_counts()
    
    return JsonResponse({'success': True})


class PostCreateView(LoginRequiredMixin, CreateView):
    model = Post
    form_class = PostForm
    template_name = 'social/post_form.html'
    success_url = reverse_lazy('social:feed')

    def form_valid(self, form):
        form.instance.author = self.request.user
        response = super().form_valid(form)
        
        # Processar hashtags
        hashtags = form.cleaned_data.get('hashtags', [])
        for hashtag_name in hashtags:
            hashtag, created = Hashtag.objects.get_or_create(name=hashtag_name)
            PostHashtag.objects.create(post=form.instance, hashtag=hashtag)
            hashtag.update_posts_count()
        
        messages.success(self.request, _('Post criado com sucesso!'))
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # Adicionar os mesmos dados do feed
        user = self.request.user
        
        # Estatísticas do usuário
        user_stats = {
            'posts_count': user.social_posts.count(),
            'followers_count': user.followers.count(),
            'following_count': user.following.count(),
            'total_likes_received': Like.objects.filter(post__author=user).count(),
            'total_comments_received': Comment.objects.filter(post__author=user).count(),
        }
        
        # Hashtags populares
        popular_hashtags = Hashtag.objects.filter(posts_count__gt=0).order_by('-posts_count')[:10]
        
        # Usuários sugeridos (usuários ativos que o usuário atual não segue)
        following_users = user.following.values_list('following_id', flat=True)
        suggested_users = User.objects.exclude(
            Q(id=user.id) | Q(id__in=following_users)
        ).filter(
            is_active=True,
            social_posts__isnull=False
        ).annotate(
            posts_count=Count('social_posts')
        ).filter(posts_count__gt=0).order_by('-posts_count')[:5]
        
        # Verificar se o usuário pode moderar
        can_moderate = user.is_superuser or user.is_staff or user.has_perm('social.can_moderate_content')
        
        # Estatísticas da rede (apenas para moderadores)
        network_stats = {}
        moderation_stats = {}
        if can_moderate:
            network_stats = {
                'total_users': User.objects.filter(is_active=True).count(),
                'total_posts': Post.objects.count(),
                'posts_today': Post.objects.filter(created_at__date=timezone.now().date()).count(),
            }
            
            moderation_stats = {
                'pending_reports': Report.objects.filter(status='pending').count(),
                'total_reports': Report.objects.count(),
                'reports_today': Report.objects.filter(created_at__date=timezone.now().date()).count(),
                'active_filters': ContentFilter.objects.filter(is_active=True).count(),
            }
        
        # Timestamp para cache busting de avatares
        import time
        timestamp = int(time.time())
        
        context.update({
            'user_stats': user_stats,
            'popular_hashtags': popular_hashtags,
            'suggested_users': suggested_users,
            'network_stats': network_stats,
            'moderation_stats': moderation_stats,
            'can_moderate': can_moderate,
            'timestamp': timestamp,
            'segment': 'post_create',
            'parent': 'social',
        })
        
        return context


class PostUpdateView(LoginRequiredMixin, UpdateView):
    model = Post
    form_class = PostForm
    template_name = 'social/post_form.html'

    def get_queryset(self):
        return Post.objects.filter(author=self.request.user)

    def get_success_url(self):
        return reverse_lazy('social:post_detail', kwargs={'post_id': self.object.id})

    def form_valid(self, form):
        # Marcar como editado
        self.object.mark_as_edited()
        
        # Processar hashtags
        hashtags = form.cleaned_data.get('hashtags', [])
        
        # Remover hashtags antigas
        self.object.hashtags.all().delete()
        
        response = super().form_valid(form)
        
        # Adicionar novas hashtags
        for hashtag_name in hashtags:
            hashtag, created = Hashtag.objects.get_or_create(name=hashtag_name)
            PostHashtag.objects.create(post=self.object, hashtag=hashtag)
            hashtag.update_posts_count()
        
        messages.success(self.request, _('Post atualizado com sucesso!'))
        return response

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['segment'] = 'post_edit'
        context['parent'] = 'social'
        
        # Pré-preencher hashtags
        if self.object:
            hashtags = [f"#{ph.hashtag.name}" for ph in self.object.hashtags.all()]
            context['form'].fields['hashtags'].initial = ' '.join(hashtags)
        
        return context


class PostDeleteView(LoginRequiredMixin, DeleteView):
    model = Post
    template_name = 'social/post_confirm_delete.html'
    success_url = reverse_lazy('social:my_posts')

    def get_queryset(self):
        return Post.objects.filter(author=self.request.user)

    def delete(self, request, *args, **kwargs):
        # Remover hashtags associadas
        post = self.get_object()
        post.hashtags.all().delete()
        
        messages.success(request, _('Post deletado com sucesso!'))
        return super().delete(request, *args, **kwargs)

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['segment'] = 'post_delete'
        context['parent'] = 'social'
        return context


# ============================================================================
# VIEWS DE MODERAÇÃO
# ============================================================================

@login_required
def report_content(request, content_type, content_id):
    """View para denunciar conteúdo"""
    if request.method == 'POST':
        form = ReportForm(request.POST)
        if form.is_valid():
            try:
                report = form.save(commit=False)
                report.reporter = request.user
                
                # Associar o conteúdo reportado
                if content_type == 'post':
                    report.reported_post = get_object_or_404(Post, id=content_id)
                elif content_type == 'comment':
                    report.reported_comment = get_object_or_404(Comment, id=content_id)
                elif content_type == 'user':
                    report.reported_user = get_object_or_404(User, id=content_id)
                else:
                    return JsonResponse({
                        'success': False,
                        'message': _('Tipo de conteúdo inválido.')
                    })
                
                report.save()
                
                # Verificar se há denúncias similares
                similar_reports = Report.objects.filter(
                    report_type=report.report_type,
                    status__in=['pending', 'reviewing']
                )
                
                if content_type == 'post':
                    similar_reports = similar_reports.filter(reported_post=report.reported_post)
                elif content_type == 'comment':
                    similar_reports = similar_reports.filter(reported_comment=report.reported_comment)
                elif content_type == 'user':
                    similar_reports = similar_reports.filter(reported_user=report.reported_user)
                
                # Atualizar contador de denúncias similares
                for similar_report in similar_reports:
                    similar_report.similar_reports_count = similar_reports.count()
                    similar_report.save()
                
                # Retornar resposta JSON para AJAX
                return JsonResponse({
                    'success': True,
                    'message': _('Denúncia enviada com sucesso. Nossa equipe irá analisar.')
                })
                
            except Exception as e:
                return JsonResponse({
                    'success': False,
                    'message': _('Erro ao processar denúncia. Tente novamente.')
                })
        else:
            # Retornar erros do formulário em JSON
            errors = {}
            for field, error_list in form.errors.items():
                errors[field] = error_list[0] if error_list else ''
            
            return JsonResponse({
                'success': False,
                'message': _('Por favor, corrija os erros no formulário.'),
                'errors': errors
            })
    
    # Para requisições GET, retornar página de denúncia (não usado no modal)
    form = ReportForm()
    
    # Obter informações do conteúdo reportado
    content_info = {}
    try:
        if content_type == 'post':
            content = get_object_or_404(Post, id=content_id)
            content_info = {
                'type': 'post',
                'content': content.content[:100] + '...' if len(content.content) > 100 else content.content,
                'author': content.author.username,
                'date': content.created_at
            }
        elif content_type == 'comment':
            content = get_object_or_404(Comment, id=content_id)
            content_info = {
                'type': 'comment',
                'content': content.content[:100] + '...' if len(content.content) > 100 else content.content,
                'author': content.author.username,
                'date': content.created_at
            }
        elif content_type == 'user':
            content = get_object_or_404(User, id=content_id)
            content_info = {
                'type': 'user',
                'content': f'Usuário: {content.username}',
                'author': content.username,
                'date': content.date_joined
            }
    except:
        return JsonResponse({
            'success': False,
            'message': _('Conteúdo não encontrado.')
        })
    
    return render(request, 'social/report_content.html', {
        'form': form,
        'content_info': content_info,
        'content_type': content_type,
        'content_id': content_id
    })


@login_required
def moderation_dashboard(request):
    """Painel principal de moderação"""
    # Verificar permissões
    if not request.user.has_perm('social.can_moderate_reports'):
        messages.error(request, _('Você não tem permissão para acessar o painel de moderação.'))
        return redirect('social:feed')
    
    # Estatísticas gerais
    total_reports = Report.objects.count()
    pending_reports = Report.objects.filter(status='pending').count()
    urgent_reports = Report.objects.filter(priority='urgent', status__in=['pending', 'reviewing']).count()
    resolved_today = Report.objects.filter(
        status='resolved',
        resolved_at__date=timezone.now().date()
    ).count()
    
    # Denúncias recentes
    recent_reports = Report.objects.filter(
        status__in=['pending', 'reviewing']
    ).select_related(
        'reporter', 'assigned_moderator', 'reported_post', 'reported_comment', 'reported_user'
    ).order_by('-priority', '-created_at')[:10]
    
    # Ações de moderação recentes
    recent_actions = ModerationAction.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).select_related('moderator', 'target_post', 'target_comment', 'target_user').order_by('-created_at')[:10]
    
    # Filtros ativos
    active_filters = ContentFilter.objects.filter(is_active=True).count()
    
    # Logs recentes
    recent_logs = ModerationLog.objects.filter(
        created_at__gte=timezone.now() - timedelta(days=7)
    ).select_related('moderator').order_by('-created_at')[:10]
    
    # Gráficos de atividade (últimos 30 dias)
    from datetime import date
    dates = []
    report_counts = []
    action_counts = []
    
    for i in range(30):
        d = date.today() - timedelta(days=i)
        dates.append(d.strftime('%d/%m'))
        report_counts.append(Report.objects.filter(created_at__date=d).count())
        action_counts.append(ModerationAction.objects.filter(created_at__date=d).count())
    
    dates.reverse()
    report_counts.reverse()
    action_counts.reverse()
    
    context = {
        'total_reports': total_reports,
        'pending_reports': pending_reports,
        'urgent_reports': urgent_reports,
        'resolved_today': resolved_today,
        'recent_reports': recent_reports,
        'recent_actions': recent_actions,
        'active_filters': active_filters,
        'recent_logs': recent_logs,
        'chart_dates': dates,
        'report_counts': report_counts,
        'action_counts': action_counts,
    }
    
    return render(request, 'social/moderation/dashboard.html', context)


@login_required
def reports_list(request):
    """Lista de denúncias para moderação"""
    if not request.user.has_perm('social.can_view_reports'):
        messages.error(request, _('Você não tem permissão para visualizar denúncias.'))
        return redirect('social:feed')
    
    # Formulário de busca
    search_form = SearchReportForm(request.GET)
    reports = Report.objects.all()
    
    if search_form.is_valid():
        q = search_form.cleaned_data.get('q')
        report_type = search_form.cleaned_data.get('report_type')
        status = search_form.cleaned_data.get('status')
        priority = search_form.cleaned_data.get('priority')
        assigned_moderator = search_form.cleaned_data.get('assigned_moderator')
        date_from = search_form.cleaned_data.get('date_from')
        date_to = search_form.cleaned_data.get('date_to')
        
        if q:
            reports = reports.filter(
                Q(description__icontains=q) |
                Q(reporter__username__icontains=q) |
                Q(reported_post__content__icontains=q) |
                Q(reported_comment__content__icontains=q) |
                Q(reported_user__username__icontains=q)
            )
        
        if report_type:
            reports = reports.filter(report_type=report_type)
        
        if status:
            reports = reports.filter(status=status)
        
        if priority:
            reports = reports.filter(priority=priority)
        
        if assigned_moderator:
            reports = reports.filter(assigned_moderator=assigned_moderator)
        
        if date_from:
            reports = reports.filter(created_at__date__gte=date_from)
        
        if date_to:
            reports = reports.filter(created_at__date__lte=date_to)
    
    # Ordenação
    order_by = request.GET.get('order_by', '-priority')
    if order_by in ['priority', '-priority', 'created_at', '-created_at', 'status']:
        reports = reports.order_by(order_by)
    else:
        reports = reports.order_by('-priority', '-created_at')
    
    # Paginação
    paginator = Paginator(reports, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Formulário para ações em massa
    bulk_form = BulkModerationForm()
    
    # Calcular estatísticas corretas
    total_reports = reports.count()
    pending_count = reports.filter(status='pending').count()
    urgent_count = reports.filter(priority='urgent', status__in=['pending', 'reviewing']).count()
    resolved_count = reports.filter(status='resolved').count()
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'bulk_form': bulk_form,
        'total_reports': total_reports,
        'pending_count': pending_count,
        'urgent_count': urgent_count,
        'resolved_count': resolved_count,
    }
    
    return render(request, 'social/moderation/reports_list.html', context)


@login_required
@require_http_methods(["POST"])
def update_report_status(request, report_id):
    """Atualiza o status de uma denúncia via AJAX"""
    if not request.user.has_perm('social.can_moderate_reports'):
        return JsonResponse({'success': False, 'error': _('Sem permissão para moderar denúncias.')}, status=403)
    
    try:
        import json
        data = json.loads(request.body)
        status = data.get('status')
        
        if status not in ['pending', 'reviewing', 'resolved', 'dismissed']:
            return JsonResponse({'success': False, 'error': _('Status inválido.')}, status=400)
        
        report = get_object_or_404(Report, id=report_id)
        old_status = report.status
        report.status = status
        report.assigned_moderator = request.user
        report.save()
        
        # Criar log de moderação
        ModerationLog.objects.create(
            moderator=request.user,
            action_type='report_status_changed',
            target_type='report',
            target_id=report.id,
            description=_('Status da denúncia #{} alterado de {} para {}').format(
                report.id, old_status, status
            )
        )
        
        return JsonResponse({
            'success': True, 
            'message': _('Status atualizado com sucesso!'),
            'new_status': status
        })
        
    except json.JSONDecodeError:
        return JsonResponse({'success': False, 'error': _('Dados inválidos.')}, status=400)
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
@require_http_methods(["POST"])
def assign_report(request, report_id):
    """Atribui uma denúncia ao moderador atual"""
    if not request.user.has_perm('social.can_moderate_reports'):
        return JsonResponse({'success': False, 'error': _('Sem permissão para moderar denúncias.')}, status=403)
    
    try:
        report = get_object_or_404(Report, id=report_id)
        
        # Verificar se já não está atribuído ao usuário atual
        if report.assigned_moderator == request.user:
            return JsonResponse({'success': False, 'error': _('Esta denúncia já está atribuída a você.')}, status=400)
        
        old_moderator = report.assigned_moderator
        report.assigned_moderator = request.user
        report.save()
        
        # Criar log de moderação
        ModerationLog.objects.create(
            moderator=request.user,
            action_type='report_assigned',
            target_type='report',
            target_id=report.id,
            description=_('Denúncia #{} atribuída a {}').format(
                report.id, request.user.username
            )
        )
        
        return JsonResponse({
            'success': True, 
            'message': _('Denúncia atribuída com sucesso!'),
            'assigned_moderator': request.user.username
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)}, status=500)


@login_required
def report_detail(request, report_id):
    """Detalhes de uma denúncia específica"""
    if not request.user.has_perm('social.can_moderate_reports'):
        messages.error(request, _('Você não tem permissão para moderar denúncias.'))
        return redirect('social:feed')
    
    report = get_object_or_404(
        Report.objects.select_related(
            'reported_post__author',
            'reported_comment__author', 
            'reported_user',
            'reporter',
            'assigned_moderator'
        ), 
        id=report_id
    )
    
    # Formulário para ação de moderação
    if request.method == 'POST':
        action_form = ModerationActionForm(request.POST)
        if action_form.is_valid():
            action = action_form.save(commit=False)
            action.moderator = request.user
            
            # Associar o alvo da ação
            if report.reported_post:
                action.target_post = report.reported_post
            elif report.reported_comment:
                action.target_comment = report.reported_comment
            elif report.reported_user:
                action.target_user = report.reported_user
            
            action.save()
            
            # Aplicar a ação
            action.apply_action()
            
            # Resolver a denúncia
            action_taken = action.get_action_type_display()
            report.resolve(request.user, action_taken, action.reason)
            
            messages.success(request, _('Ação de moderação aplicada com sucesso.'))
            return redirect('social:reports_list')
    else:
        action_form = ModerationActionForm()
    
    # Denúncias similares
    similar_reports = Report.objects.filter(
        report_type=report.report_type,
        status__in=['pending', 'reviewing']
    ).exclude(id=report.id)
    
    if report.reported_post:
        similar_reports = similar_reports.filter(reported_post=report.reported_post)
    elif report.reported_comment:
        similar_reports = similar_reports.filter(reported_comment=report.reported_comment)
    elif report.reported_user:
        similar_reports = similar_reports.filter(reported_user=report.reported_user)
    
    # Histórico de ações do usuário reportado
    user_actions = []
    if report.reported_user:
        user_actions = ModerationAction.objects.filter(
            target_user=report.reported_user
        ).order_by('-created_at')[:10]
    
    context = {
        'report': report,
        'action_form': action_form,
        'similar_reports': similar_reports,
        'user_actions': user_actions,
    }
    
    return render(request, 'social/moderation/report_detail.html', context)


@login_required
def content_filters(request):
    """Gerenciamento de filtros de conteúdo"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        messages.error(request, _('Você não tem permissão para gerenciar filtros.'))
        return redirect('social:feed')
    
    filters = ContentFilter.objects.all().order_by('-is_active', 'name')
    
    if request.method == 'POST':
        filter_form = ContentFilterForm(request.POST)
        if filter_form.is_valid():
            content_filter = filter_form.save()
            messages.success(request, _('Filtro criado com sucesso.'))
            return redirect('social:content_filters')
    else:
        filter_form = ContentFilterForm()
    
    # Calcular estatísticas dos filtros
    total_matches = sum(filter.matches_count for filter in filters)
    last_activity = None
    if filters.exists():
        last_matched_filter = filters.exclude(last_matched__isnull=True).order_by('-last_matched').first()
        if last_matched_filter:
            last_activity = last_matched_filter.last_matched
    
    context = {
        'filters': filters,
        'filter_form': filter_form,
        'total_filters': filters.count(),
        'active_filters': filters.filter(is_active=True).count(),
        'total_matches': total_matches,
        'last_activity': last_activity,
    }
    
    return render(request, 'social/moderation/content_filters.html', context)


@login_required
def edit_filter(request, filter_id):
    """Editar filtro de conteúdo"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        messages.error(request, _('Você não tem permissão para editar filtros.'))
        return redirect('social:feed')
    
    content_filter = get_object_or_404(ContentFilter, id=filter_id)
    
    if request.method == 'POST':
        filter_form = ContentFilterForm(request.POST, instance=content_filter)
        if filter_form.is_valid():
            filter_form.save()
            messages.success(request, _('Filtro atualizado com sucesso.'))
            return redirect('social:content_filters')
    else:
        filter_form = ContentFilterForm(instance=content_filter)
    
    context = {
        'filter_form': filter_form,
        'content_filter': content_filter,
    }
    
    return render(request, 'social/moderation/edit_filter.html', context)


@login_required
def toggle_filter(request, filter_id):
    """Ativar/desativar filtro"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        return JsonResponse({'error': _('Permissão negada')}, status=403)
    
    content_filter = get_object_or_404(ContentFilter, id=filter_id)
    content_filter.is_active = not content_filter.is_active
    content_filter.save()
    
    return JsonResponse({
        'success': True,
        'is_active': content_filter.is_active,
        'message': _('Filtro ativado') if content_filter.is_active else _('Filtro desativado')
    })


@login_required
def delete_filter(request, filter_id):
    """Deletar filtro de conteúdo"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        return JsonResponse({'error': _('Permissão negada')}, status=403)
    
    if request.method != 'POST':
        return JsonResponse({'error': _('Método não permitido')}, status=405)
    
    content_filter = get_object_or_404(ContentFilter, id=filter_id)
    
    # Log da ação
    ModerationLog.log_action(
        moderator=request.user,
        action_type='filter_deleted',
        target_type='filter',
        target_id=filter_id,
        description=f"Filtro '{content_filter.name}' foi deletado",
        details=f"Tipo: {content_filter.get_filter_type_display()}\nPadrão: {content_filter.pattern[:100]}..."
    )
    
    content_filter.delete()
    
    return JsonResponse({
        'success': True,
        'message': _('Filtro deletado com sucesso')
    })


@login_required
def moderation_logs(request):
    """Logs de moderação"""
    if not request.user.has_perm('social.can_view_moderation_logs'):
        messages.error(request, _('Você não tem permissão para visualizar logs.'))
        return redirect('social:feed')
    
    logs = ModerationLog.objects.all().select_related('moderator').order_by('-created_at')
    
    # Filtros
    action_type = request.GET.get('action_type')
    moderator_id = request.GET.get('moderator')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    
    if moderator_id:
        logs = logs.filter(moderator_id=moderator_id)
    
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    # Paginação
    paginator = Paginator(logs, 50)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Estatísticas
    total_actions = logs.count()
    actions_today = logs.filter(created_at__date=timezone.now().date()).count()
    
    # Moderadores ativos
    active_moderators = User.objects.filter(
        moderation_logs__created_at__gte=timezone.now() - timedelta(days=30)
    ).distinct().count()
    
    context = {
        'page_obj': page_obj,
        'total_actions': total_actions,
        'actions_today': actions_today,
        'active_moderators': active_moderators,
        'action_types': ModerationLog.LOG_TYPES,
        'moderators': User.objects.filter(is_staff=True),
        'current_filters': {
            'action_type': action_type,
            'moderator_id': moderator_id,
            'date_from': date_from,
            'date_to': date_to,
        }
    }
    
    return render(request, 'social/moderation/logs.html', context)


@login_required
def export_logs_excel(request):
    """Exportar logs de moderação em Excel formatado"""
    if not request.user.has_perm('social.can_view_moderation_logs'):
        messages.error(request, _('Você não tem permissão para exportar logs.'))
        return redirect('social:moderation_logs')
    
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, _('Biblioteca openpyxl não está instalada. Use: pip install openpyxl'))
        return redirect('social:moderation_logs')
    
    from django.http import HttpResponse
    from datetime import datetime
    
    # Aplicar os mesmos filtros da view de logs
    logs = ModerationLog.objects.all().select_related('moderator').order_by('-created_at')
    
    # Filtros da query string
    action_type = request.GET.get('action_type')
    moderator_id = request.GET.get('moderator')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    if moderator_id:
        logs = logs.filter(moderator_id=moderator_id)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    # Limitar a 10.000 registros para evitar problemas de performance
    logs = logs[:10000]
    
    # Criar workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Logs de Moderação"
    
    # Estilos
    header_font = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    header_alignment = Alignment(horizontal='center', vertical='center')
    
    border_thin = Border(
        left=Side(style='thin', color='D1D1D1'),
        right=Side(style='thin', color='D1D1D1'),
        top=Side(style='thin', color='D1D1D1'),
        bottom=Side(style='thin', color='D1D1D1')
    )
    
    data_font = Font(name='Calibri', size=10)
    data_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    # Cabeçalhos
    headers = [
        'Data/Hora', 'Moderador', 'Tipo de Ação', 'Tipo do Alvo', 
        'ID do Alvo', 'Descrição', 'Detalhes', 'IP', 'Navegador'
    ]
    
    # Configurar cabeçalho
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border_thin
    
    # Dados
    for row_num, log in enumerate(logs, 2):
        # Data formatada brasileiro
        ws.cell(row=row_num, column=1, value=log.created_at.strftime('%d/%m/%Y %H:%M:%S'))
        
        # Moderador com nome completo ou username
        moderator_name = log.moderator.get_full_name() if log.moderator and log.moderator.get_full_name() else (
            log.moderator.username if log.moderator else 'Sistema Automático'
        )
        ws.cell(row=row_num, column=2, value=moderator_name)
        
        # Tipo de ação traduzido
        ws.cell(row=row_num, column=3, value=log.get_action_type_display())
        
        # Tipo do alvo com primeira letra maiúscula
        ws.cell(row=row_num, column=4, value=log.target_type.title())
        
        # ID do alvo
        ws.cell(row=row_num, column=5, value=log.target_id)
        
        # Descrição limitada
        description = log.description[:300] + '...' if len(log.description) > 300 else log.description
        ws.cell(row=row_num, column=6, value=description)
        
        # Detalhes limitados
        details = log.details
        if details:
            details = details[:300] + '...' if len(details) > 300 else details
        else:
            details = '-'
        ws.cell(row=row_num, column=7, value=details)
        
        # IP
        ws.cell(row=row_num, column=8, value=log.ip_address or '-')
        
        # Navegador simplificado
        user_agent = '-'
        if log.user_agent:
            ua = log.user_agent.lower()
            if 'chrome' in ua and 'edge' not in ua:
                user_agent = 'Chrome'
            elif 'firefox' in ua:
                user_agent = 'Firefox'
            elif 'safari' in ua and 'chrome' not in ua:
                user_agent = 'Safari'
            elif 'edge' in ua:
                user_agent = 'Edge'
            elif 'opera' in ua:
                user_agent = 'Opera'
            else:
                user_agent = 'Outro'
        ws.cell(row=row_num, column=9, value=user_agent)
        
        # Aplicar estilos
        for col_num in range(1, 10):
            cell = ws.cell(row=row_num, column=col_num)
            cell.font = data_font
            cell.alignment = data_alignment
            cell.border = border_thin
    
    # Ajustar largura das colunas
    column_widths = {
        1: 16,  # Data/Hora
        2: 18,  # Moderador  
        3: 22,  # Tipo de Ação
        4: 14,  # Tipo do Alvo
        5: 8,   # ID do Alvo
        6: 45,  # Descrição
        7: 35,  # Detalhes
        8: 14,  # IP
        9: 12,  # Navegador
    }
    
    for col_num, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_num)].width = width
    
    # Adicionar aba de estatísticas e filtros
    stats_ws = wb.create_sheet(title="Relatório e Estatísticas")
    
    # Cabeçalho do relatório
    title_font = Font(name='Calibri', size=16, bold=True, color='2E75B6')
    stats_ws.cell(row=1, column=1, value="RELATÓRIO DE LOGS DE MODERAÇÃO").font = title_font
    stats_ws.cell(row=2, column=1, value=f"Gerado em: {datetime.now().strftime('%d/%m/%Y às %H:%M:%S')}")
    stats_ws.cell(row=3, column=1, value=f"Total de registros: {logs.count():,}")
    
    # Filtros aplicados
    row = 5
    if any([action_type, moderator_id, date_from, date_to]):
        stats_ws.cell(row=row, column=1, value="FILTROS APLICADOS:").font = Font(bold=True)
        row += 1
        if action_type:
            action_display = dict(ModerationLog.LOG_TYPES).get(action_type, action_type)
            stats_ws.cell(row=row, column=1, value=f"• Tipo de Ação: {action_display}")
            row += 1
        if moderator_id:
            try:
                moderator = User.objects.get(id=moderator_id)
                moderator_name = moderator.get_full_name() or moderator.username
                stats_ws.cell(row=row, column=1, value=f"• Moderador: {moderator_name}")
                row += 1
            except User.DoesNotExist:
                pass
        if date_from:
            stats_ws.cell(row=row, column=1, value=f"• Data inicial: {date_from}")
            row += 1
        if date_to:
            stats_ws.cell(row=row, column=1, value=f"• Data final: {date_to}")
            row += 1
        row += 1
    
    # Estatísticas por tipo de ação
    stats_ws.cell(row=row, column=1, value="AÇÕES POR TIPO:").font = Font(bold=True)
    stats_ws.cell(row=row, column=2, value="Quantidade").font = Font(bold=True)
    row += 1
    
    action_stats = {}
    for log in logs:
        action_type_display = log.get_action_type_display()
        action_stats[action_type_display] = action_stats.get(action_type_display, 0) + 1
    
    for action, count in sorted(action_stats.items(), key=lambda x: x[1], reverse=True):
        stats_ws.cell(row=row, column=1, value=action)
        stats_ws.cell(row=row, column=2, value=count)
        row += 1
    
    row += 1
    
    # Estatísticas por moderador
    stats_ws.cell(row=row, column=1, value="AÇÕES POR MODERADOR:").font = Font(bold=True)
    stats_ws.cell(row=row, column=2, value="Quantidade").font = Font(bold=True)
    row += 1
    
    moderator_stats = {}
    for log in logs:
        moderator_name = log.moderator.get_full_name() if log.moderator and log.moderator.get_full_name() else (
            log.moderator.username if log.moderator else 'Sistema Automático'
        )
        moderator_stats[moderator_name] = moderator_stats.get(moderator_name, 0) + 1
    
    for moderator, count in sorted(moderator_stats.items(), key=lambda x: x[1], reverse=True):
        stats_ws.cell(row=row, column=1, value=moderator)
        stats_ws.cell(row=row, column=2, value=count)
        row += 1
    
    # Ajustar largura das colunas da aba de estatísticas
    stats_ws.column_dimensions['A'].width = 35
    stats_ws.column_dimensions['B'].width = 12
    
    # Preparar resposta
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    
    # Nome do arquivo com timestamp e filtros
    filename_parts = ['logs_moderacao']
    if action_type:
        filename_parts.append(f'acao_{action_type}')
    if date_from or date_to:
        if date_from and date_to:
            filename_parts.append(f'{date_from}_a_{date_to}')
        elif date_from:
            filename_parts.append(f'desde_{date_from}')
        elif date_to:
            filename_parts.append(f'ate_{date_to}')
    
    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(filename_parts) + '.xlsx'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    wb.save(response)
    return response


@login_required
def export_logs_csv(request):
    """Exportar logs de moderação em CSV"""
    if not request.user.has_perm('social.can_view_moderation_logs'):
        messages.error(request, _('Você não tem permissão para exportar logs.'))
        return redirect('social:moderation_logs')
    
    import csv
    from django.http import HttpResponse
    from datetime import datetime
    
    # Aplicar os mesmos filtros da view de logs
    logs = ModerationLog.objects.all().select_related('moderator').order_by('-created_at')
    
    action_type = request.GET.get('action_type')
    moderator_id = request.GET.get('moderator')
    date_from = request.GET.get('date_from')
    date_to = request.GET.get('date_to')
    
    if action_type:
        logs = logs.filter(action_type=action_type)
    if moderator_id:
        logs = logs.filter(moderator_id=moderator_id)
    if date_from:
        logs = logs.filter(created_at__date__gte=date_from)
    if date_to:
        logs = logs.filter(created_at__date__lte=date_to)
    
    # Preparar resposta CSV
    response = HttpResponse(content_type='text/csv; charset=utf-8')
    
    filename_parts = ['logs_moderacao']
    if action_type:
        filename_parts.append(f'acao_{action_type}')
    if date_from or date_to:
        if date_from and date_to:
            filename_parts.append(f'{date_from}_a_{date_to}')
        elif date_from:
            filename_parts.append(f'desde_{date_from}')
        elif date_to:
            filename_parts.append(f'ate_{date_to}')
    
    filename_parts.append(datetime.now().strftime('%Y%m%d_%H%M%S'))
    filename = '_'.join(filename_parts) + '.csv'
    
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # BOM para UTF-8 (para Excel abrir corretamente)
    response.write('\ufeff')
    
    writer = csv.writer(response)
    writer.writerow([
        'Data/Hora', 'Moderador', 'Tipo de Ação', 'Tipo do Alvo', 'ID do Alvo',
        'Descrição', 'Detalhes', 'Endereço IP', 'Navegador'
    ])
    
    for log in logs:
        moderator_name = log.moderator.get_full_name() if log.moderator and log.moderator.get_full_name() else (
            log.moderator.username if log.moderator else 'Sistema Automático'
        )
        
        # Navegador simplificado
        user_agent = '-'
        if log.user_agent:
            ua = log.user_agent.lower()
            if 'chrome' in ua and 'edge' not in ua:
                user_agent = 'Chrome'
            elif 'firefox' in ua:
                user_agent = 'Firefox'
            elif 'safari' in ua and 'chrome' not in ua:
                user_agent = 'Safari'
            elif 'edge' in ua:
                user_agent = 'Edge'
            elif 'opera' in ua:
                user_agent = 'Opera'
            else:
                user_agent = 'Outro'
        
        writer.writerow([
            log.created_at.strftime('%d/%m/%Y %H:%M:%S'),
            moderator_name,
            log.get_action_type_display(),
            log.target_type.title(),
            log.target_id,
            log.description,
            log.details or '-',
            log.ip_address or '-',
            user_agent
        ])
    
    return response


@login_required
def apply_retroactive_filters(request):
    """Aplicar filtros a conteúdo retroativo"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        messages.error(request, _('Você não tem permissão para esta ação.'))
        return redirect('social:moderation_dashboard')
    
    if request.method == 'POST':
        from django.core.management import call_command
        from io import StringIO
        import json
        
        content_type = request.POST.get('content_type', 'all')
        dry_run = request.POST.get('dry_run') == 'on'
        filter_id = request.POST.get('filter_id')
        
        try:
            # Capturar saída do comando
            out = StringIO()
            
            # Preparar argumentos do comando
            command_args = [
                '--content-type', content_type,
                '--batch-size', '50',  # Lotes menores para interface web
            ]
            
            if dry_run:
                command_args.append('--dry-run')
            
            if filter_id:
                command_args.extend(['--filter-id', filter_id])
            
            # Executar comando
            call_command('apply_filters_retroactive', *command_args, stdout=out)
            
            # Capturar resultado
            output = out.getvalue()
            
            # Log da ação
            ModerationLog.log_action(
                moderator=request.user,
                action_type='report_created',
                target_type='system',
                target_id=0,
                description=f"Filtros retroativos aplicados via interface web",
                details=f"Tipo: {content_type}, Dry-run: {dry_run}, Filtro: {filter_id or 'todos'}",
                request=request
            )
            
            if dry_run:
                messages.info(request, _('Simulação concluída! Verifique os logs para detalhes.'))
            else:
                messages.success(request, _('Filtros aplicados com sucesso ao conteúdo existente!'))
            
            # Retornar resultado para AJAX
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': True,
                    'message': _('Operação concluída com sucesso'),
                    'output': output,
                    'dry_run': dry_run
                })
            
        except Exception as e:
            error_msg = f'Erro ao aplicar filtros retroativos: {str(e)}'
            messages.error(request, error_msg)
            
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_msg
                })
    
    # Obter filtros disponíveis para o formulário
    active_filters = ContentFilter.objects.filter(is_active=True).order_by('name')
    
    context = {
        'active_filters': active_filters,
        'segment': 'moderation',
        'parent': 'social',
    }
    
    return render(request, 'social/moderation/apply_retroactive.html', context)


@login_required
def bulk_moderation_action(request):
    """Ação em massa de moderação"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        return JsonResponse({'error': _('Permissão negada')}, status=403)
    
    if request.method == 'POST':
        report_ids = request.POST.getlist('report_ids')
        action_type = request.POST.get('action_type')
        reason = request.POST.get('reason', '')
        assigned_moderator_id = request.POST.get('assigned_moderator')
        
        if not report_ids or not action_type:
            return JsonResponse({'error': _('Dados inválidos')}, status=400)
        
        reports = Report.objects.filter(id__in=report_ids)
        processed_count = 0
        failed_count = 0
        errors = []
        
        if not reports.exists():
            return JsonResponse({'error': _('Nenhuma denúncia válida encontrada')}, status=400)
        
        for report in reports:
            try:
                if action_type == 'assign_moderator' and assigned_moderator_id:
                    # Verificar se o moderador existe e tem permissões
                    try:
                        moderator = get_user_model().objects.get(id=assigned_moderator_id, is_staff=True)
                        if not moderator.has_perm('social.can_take_moderation_actions'):
                            raise ValueError(_('Moderador selecionado não tem permissões necessárias'))
                    except get_user_model().DoesNotExist:
                        raise ValueError(_('Moderador selecionado não encontrado'))
                    
                    report.assigned_moderator_id = assigned_moderator_id
                    report.status = 'reviewing'
                    report.save()
                    processed_count += 1
                    
                    # Log da ação
                    try:
                        ModerationLog.log_action(
                            moderator=request.user,
                            action_type='report_assigned',
                            target_type='report',
                            target_id=report.id,
                            description=f'Denúncia atribuída ao moderador',
                            details=reason
                        )
                    except Exception:
                        # Falha no log não deve impedir a operação principal
                        pass
                
                elif action_type in ['hide_content', 'delete_content', 'warn']:
                    # Verificar se há conteúdo para aplicar a ação
                    if not (report.reported_post or report.reported_comment or report.reported_user):
                        raise ValueError(_('Denúncia não possui conteúdo válido para aplicar ação'))
                    
                    # Validações específicas por tipo de ação
                    if action_type in ['hide_content', 'delete_content']:
                        # Verificar se o conteúdo ainda existe
                        content_exists = False
                        if report.reported_post:
                            try:
                                # Verificar se o post ainda existe
                                Post.objects.get(id=report.reported_post.id)
                                content_exists = True
                            except Post.DoesNotExist:
                                raise ValueError(_('Post reportado não existe mais'))
                        elif report.reported_comment:
                            try:
                                # Verificar se o comentário ainda existe
                                Comment.objects.get(id=report.reported_comment.id)
                                content_exists = True
                            except Comment.DoesNotExist:
                                raise ValueError(_('Comentário reportado não existe mais'))
                        
                        if not content_exists:
                            raise ValueError(_('Conteúdo reportado não existe mais'))
                    
                    if action_type == 'warn' and not report.reported_user:
                        raise ValueError(_('Ação de advertência requer usuário reportado'))
                    
                    # Criar ação de moderação
                    action = ModerationAction.objects.create(
                        moderator=request.user,
                        action_type=action_type,
                        reason=reason,
                        target_post=report.reported_post,
                        target_comment=report.reported_comment,
                        target_user=report.reported_user
                    )
                    
                    # Aplicar ação
                    action.apply_action()
                    
                    # Resolver denúncia
                    report.resolve(request.user, action.get_action_type_display(), reason)
                    processed_count += 1
                    
            except Exception as e:
                failed_count += 1
                error_msg = f'Denúncia #{report.id}: {str(e)}'
                errors.append(error_msg)
                
                # Log do erro mas continua processando outras denúncias
                try:
                    ModerationLog.log_action(
                        moderator=request.user,
                        action_type='bulk_action_error',
                        target_type='report',
                        target_id=report.id,
                        description=f'Erro ao processar denúncia: {str(e)}',
                        details=reason
                    )
                except Exception:
                    # Falha no log não deve impedir a operação principal
                    pass
        
        # Preparar resposta
        success_message = _('Ação aplicada com sucesso')
        if failed_count > 0:
            success_message = _('Ação aplicada com sucesso em %(processed)d denúncia(s). %(failed)d falharam.') % {
                'processed': processed_count,
                'failed': failed_count
            }
        
        return JsonResponse({
            'success': True,
            'processed_count': processed_count,
            'failed_count': failed_count,
            'total_count': len(report_ids),
            'errors': errors,
            'message': success_message
        })
    
    return JsonResponse({'error': _('Método não permitido')}, status=405)


@login_required
def test_content_filter(request):
    """Testar filtro de conteúdo"""
    if not request.user.has_perm('social.can_take_moderation_actions'):
        messages.error(request, _('Você não tem permissão para testar filtros.'))
        return redirect('social:content_filters')
    
    if request.method == 'POST':
        pattern = request.POST.get('pattern', '').strip()
        filter_type = request.POST.get('filter_type', 'keyword')
        content = request.POST.get('content', '').strip()
        
        if not pattern or not content:
            messages.error(request, _('Padrão e conteúdo são obrigatórios.'))
            return redirect('social:content_filters')
        
        # Criar um filtro temporário para teste (não salva no banco)
        temp_filter = ContentFilter(
            name='Teste Temporário',
            filter_type=filter_type,
            pattern=pattern,
            is_active=True,
            case_sensitive=False
        )
        
        # Testar se o conteúdo corresponde ao filtro
        matches = temp_filter.matches_content(content)
        
        # Redirecionar com resultado
        filters = ContentFilter.objects.all().order_by('-is_active', 'name')
        
        # Calcular estatísticas dos filtros
        total_matches = sum(filter.matches_count for filter in filters)
        last_activity = None
        if filters.exists():
            last_matched_filter = filters.exclude(last_matched__isnull=True).order_by('-last_matched').first()
            if last_matched_filter:
                last_activity = last_matched_filter.last_matched
        
        context = {
            'filters': filters,
            'filter_form': ContentFilterForm(),
            'total_filters': filters.count(),
            'active_filters': filters.filter(is_active=True).count(),
            'total_matches': total_matches,
            'last_activity': last_activity,
            'test_result': matches,
            'test_pattern': pattern,
            'test_filter_type': filter_type,
            'test_content': content,
        }
        
        return render(request, 'social/moderation/content_filters.html', context)
    
    return redirect('social:content_filters')
